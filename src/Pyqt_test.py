import sys

import openpyxl



class BaseExcel(object):
    def __init__(self, file_path, sheet_name, data_o=False, read_o=False, kp_vb=False, kp_lk=False):
        self.file_name = file_path.split('\\')[-1]
        self.sheet_name = sheet_name
        self.wb = openpyxl.load_workbook(filename=self.file_name, data_only=data_o, read_only=read_o, keep_vba=kp_vb,
                                         keep_links=kp_lk)
        self.ws = self.wb[sheet_name]

    def enter_data_by_vertical(self, keyword, data):
        key_coordinate = None
        not_found_text = None
        for rows in self.ws.iter_rows():
            for cell in rows:
                if cell.value == keyword:
                    key_coordinate = cell.coordinate
                    for count in range(cell.row + 1, self.ws.max_row + 1):
                        self.ws['%s%d' % (cell.column, count)].value = data
                else:
                    not_found_text = '请查询是否有关键字'
        self.wb.save(filename=self.file_name)
        self.wb.close()
        return key_coordinate, not_found_text

    def enter_data_by_vertical_limit(self, keyword, data, limit_row):
        key_coordinate = None
        not_found_text = None
        for rows in self.ws.iter_rows():
            for cell in rows:
                if cell.value == keyword:
                    key_coordinate = cell.coordinate
                    for count in range(cell.row + 1, limit_row + 1):
                        self.ws['%s%d' % (cell.column, count)].value = data
                else:
                    not_found_text = '请查询是否有关键字'
        self.wb.save(filename=self.file_name)
        self.wb.close()
        return key_coordinate, not_found_text


